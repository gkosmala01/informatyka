### Grzegorz Kosmala; 49564
### 18.01.2021 r.
### zadanie B18
### plik "Zeszyt1.xlsx" znajduje się w repozytorium

from openpyxl import load_workbook

wb = load_workbook('Zeszyt1.xlsx')

sheet = wb.active

for row in sheet.iter_rows(min_row=1, min_col=1, max_row=10, max_col=3):
    for cell in row:
        print(cell.value)

niepusta_komorka = 0
for row in sheet.iter_rows(min_row=1, min_col=1,max_row=10, max_col=3):
    for cell in row:
        if cell.value != 0:
             print(cell.column,'-',cell.row, '---> niepusta komorka')
